from __future__ import annotations

import argparse
import base64
import csv
import fnmatch
import hashlib
import hmac
import json
import mimetypes
import os
import re
import secrets
import signal
import subprocess
import sys
import threading
import time
import uuid
import webbrowser
from copy import deepcopy
from datetime import datetime
from http import HTTPStatus
from http.cookies import SimpleCookie
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook


APP_VERSION = "4.0.0"
PROJECT_ROOT = Path(__file__).resolve().parent
DATA_ROOT = Path(os.environ.get("ITCYBER_DATA_DIR", PROJECT_ROOT / "dashboard_data")).resolve()
OUTPUT_ROOT = Path(os.environ.get("ITCYBER_OUTPUT_DIR", PROJECT_ROOT / "output")).resolve()
HISTORY_FILE = DATA_ROOT / "jobs.json"
ADMIN_FILE = DATA_ROOT / "admin.json"
ACTIVE_STATUSES = {"queued", "running", "stopping"}
MAX_LOG_LINES = 800
SESSION_COOKIE = "itcyber_admin_session"
SESSION_SECONDS = 8 * 60 * 60
PASSWORD_ITERATIONS = 390_000

# Direct-access build: no login is required.
AUTH_DISABLED = True
OPEN_DASHBOARD_USER = {"name": "Admin", "email": "direct-access"}


DEFAULT_ALLOWED_ORIGINS = (
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "http://localhost:5500",
    "http://127.0.0.1:5500",
    "https://*.vercel.app",
)


def configured_origins() -> tuple[str, ...]:
    """
    Allowed frontend origins.

    Production:
      Set ITCYBER_ALLOWED_ORIGINS to your exact Vercel URL, for example:
      https://your-project.vercel.app

    Multiple origins can be comma-separated.
    Wildcards such as https://*.vercel.app are supported.
    """
    raw = os.environ.get("ITCYBER_ALLOWED_ORIGINS", "").strip()
    if raw:
        origins = raw.split(",")
    else:
        origins = DEFAULT_ALLOWED_ORIGINS
    return tuple(str(origin).strip().rstrip("/") for origin in origins if str(origin).strip())


def origin_is_allowed(origin: str) -> bool:
    normalized = str(origin or "").strip().rstrip("/")
    if not normalized:
        return True
    return any(fnmatch.fnmatchcase(normalized, pattern) for pattern in configured_origins())


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def normalized_email(value: str) -> str:
    email = str(value or "").strip().casefold()
    if len(email) > 254 or not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", email):
        raise ValueError("Enter a valid admin email address.")
    return email


def validate_password(password: str):
    if len(password) < 8:
        raise ValueError("Password must contain at least 8 characters.")
    if len(password) > 256:
        raise ValueError("Password is too long.")
    if not re.search(r"[A-Za-z]", password) or not re.search(r"\d", password):
        raise ValueError("Password must contain at least one letter and one number.")


def password_record(password: str, salt: bytes | None = None) -> dict:
    validate_password(password)
    salt = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, PASSWORD_ITERATIONS)
    return {
        "algorithm": "pbkdf2_sha256", "iterations": PASSWORD_ITERATIONS,
        "salt": base64.b64encode(salt).decode("ascii"),
        "password_hash": base64.b64encode(digest).decode("ascii"),
    }


def password_matches(password: str, record: dict) -> bool:
    try:
        salt = base64.b64decode(record["salt"], validate=True)
        expected = base64.b64decode(record["password_hash"], validate=True)
        iterations = int(record.get("iterations", PASSWORD_ITERATIONS))
        actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
        return hmac.compare_digest(actual, expected)
    except (KeyError, TypeError, ValueError):
        return False


class AuthManager:
    def __init__(self, config_path: Path = ADMIN_FILE, bootstrap_from_env: bool = True):
        self.config_path = Path(config_path)
        self.lock = threading.RLock()
        self.config: dict | None = None
        self.sessions: dict[str, dict] = {}
        self.failures: dict[str, dict] = {}
        self._load()
        if bootstrap_from_env:
            self._bootstrap_from_environment()

    def _bootstrap_from_environment(self):
        if self.config is not None:
            return
        values = {
            "name": os.environ.get("ITCYBER_BOOTSTRAP_ADMIN_NAME", "").strip(),
            "email": os.environ.get("ITCYBER_BOOTSTRAP_ADMIN_EMAIL", "").strip(),
            "password": os.environ.get("ITCYBER_BOOTSTRAP_ADMIN_PASSWORD", ""),
        }
        if not any(values.values()):
            return
        if not all(values.values()):
            # Do not crash the whole Railway service just because one optional
            # bootstrap variable is missing. The admin can still be created
            # through /api/auth/setup after deployment.
            missing = [key for key, value in values.items() if not value]
            print(
                "WARNING: incomplete admin bootstrap variables; setup will remain available. "
                f"Missing: {', '.join(missing)}",
                file=sys.stderr,
                flush=True,
            )
            return
        token, _ = self.setup(values["name"], values["email"], values["password"])
        self.logout(token)

    def _load(self):
        if not self.config_path.exists():
            return
        try:
            value = json.loads(self.config_path.read_text(encoding="utf-8"))
            if isinstance(value, dict) and value.get("email") and value.get("password_hash"):
                self.config = value
        except (OSError, json.JSONDecodeError):
            self.config = None

    @property
    def setup_required(self) -> bool:
        return self.config is None

    def _save(self):
        self.config_path.parent.mkdir(parents=True, exist_ok=True)
        temporary = self.config_path.with_suffix(".tmp")
        temporary.write_text(json.dumps(self.config, ensure_ascii=False, indent=2), encoding="utf-8")
        try:
            os.chmod(temporary, 0o600)
        except OSError:
            pass
        temporary.replace(self.config_path)

    def _public_user(self) -> dict | None:
        if not self.config:
            return None
        return {"name": self.config.get("name", "Admin"), "email": self.config["email"]}

    def setup(self, name: str, email: str, password: str) -> tuple[str, dict]:
        name = " ".join(str(name or "").split())
        if len(name) < 2 or len(name) > 80:
            raise ValueError("Admin name must contain 2 to 80 characters.")
        email = normalized_email(email)
        record = password_record(password)
        with self.lock:
            if not self.setup_required:
                raise RuntimeError("Admin account is already configured.")
            self.config = {"name": name, "email": email, **record, "created_at": now_iso(), "updated_at": now_iso()}
            self._save()
            return self._new_session()

    def _prune_sessions(self):
        now = time.time()
        self.sessions = {token: session for token, session in self.sessions.items() if session["expires_at"] > now}

    def _new_session(self) -> tuple[str, dict]:
        self._prune_sessions()
        token = secrets.token_urlsafe(40)
        self.sessions[token] = {"expires_at": time.time() + SESSION_SECONDS, "user": self._public_user()}
        return token, deepcopy(self.sessions[token]["user"])

    def login(self, email: str, password: str, client_key: str) -> tuple[str, dict]:
        email = str(email or "").strip().casefold()
        password = str(password or "")
        now = time.time()
        with self.lock:
            attempt = self.failures.get(client_key, {"count": 0, "first_at": now, "locked_until": 0})
            if attempt.get("locked_until", 0) > now:
                wait = max(1, round(attempt["locked_until"] - now))
                raise PermissionError(f"Too many login attempts. Try again in {wait} seconds.")
            valid = bool(
                self.config and hmac.compare_digest(email, self.config["email"])
                and password_matches(password, self.config)
            )
            if not valid:
                if now - attempt.get("first_at", now) > 10 * 60:
                    attempt = {"count": 0, "first_at": now, "locked_until": 0}
                attempt["count"] = attempt.get("count", 0) + 1
                if attempt["count"] >= 5:
                    attempt["locked_until"] = now + 60
                    attempt["count"] = 0
                    attempt["first_at"] = now
                self.failures[client_key] = attempt
                raise PermissionError("Invalid admin email or password.")
            self.failures.pop(client_key, None)
            return self._new_session()

    def session_user(self, token: str | None) -> dict | None:
        if not token:
            return None
        with self.lock:
            self._prune_sessions()
            session = self.sessions.get(token)
            return deepcopy(session["user"]) if session else None

    def logout(self, token: str | None):
        if token:
            with self.lock:
                self.sessions.pop(token, None)

    def change_password(self, token: str, current_password: str, new_password: str):
        with self.lock:
            if not self.session_user(token):
                raise PermissionError("Admin session expired. Please sign in again.")
            if not self.config or not password_matches(current_password, self.config):
                raise PermissionError("Current password is incorrect.")
            self.config.update(password_record(new_password))
            self.config["updated_at"] = now_iso()
            self._save()
            current = self.sessions.get(token)
            self.sessions.clear()
            if current:
                current["expires_at"] = time.time() + SESSION_SECONDS
                self.sessions[token] = current


def clamp_int(value, minimum: int, maximum: int, field: str) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field} must be a whole number.") from exc
    if not minimum <= number <= maximum:
        raise ValueError(f"{field} must be between {minimum} and {maximum}.")
    return number


def clamp_float(value, minimum: float, maximum: float, field: str) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field} must be a number.") from exc
    if not minimum <= number <= maximum:
        raise ValueError(f"{field} must be between {minimum} and {maximum}.")
    return number


def validate_job_payload(payload: dict) -> dict:
    query = " ".join(str(payload.get("query", "")).split())
    if not query:
        raise ValueError("Enter a business category and location.")
    if len(query) > 200:
        raise ValueError("Query must be 200 characters or fewer.")
    enrichment = str(payload.get("enrichment", "website"))
    if enrichment not in {"none", "website", "full"}:
        raise ValueError("Invalid enrichment mode.")
    output_format = str(payload.get("format", "xlsx"))
    if output_format not in {"xlsx", "csv", "json"}:
        raise ValueError("Invalid output format.")
    return {
        "query": query,
        "max_results": clamp_int(payload.get("max_results", 100), 1, 2000, "Maximum results"),
        "max_website_pages": clamp_int(payload.get("max_website_pages", 2), 1, 8, "Website pages"),
        "delay": clamp_float(payload.get("delay", 2.5), 2.0, 30.0, "Delay"),
        "enrichment": enrichment,
        "format": output_format,
    }


def parse_progress_line(line: str, progress: dict) -> dict:
    updated = dict(progress)
    match = re.search(r"Discovery progress:\s*(\d+)/(\d+)", line)
    if match:
        updated.update(stage="discovering", discovered=int(match.group(1)), discovery_target=int(match.group(2)))
    match = re.search(r"\[(\d+)/(\d+)\]\s*Reading Google listing", line)
    if match:
        updated.update(stage="extracting", current=int(match.group(1)), total=int(match.group(2)))
    match = re.search(r"Checkpoint saved \((\d+) record", line)
    if match:
        updated.update(stage="enriching", records=int(match.group(1)))
    match = re.search(r"Finished\. Saved (\d+) record", line)
    if match:
        updated.update(stage="completed", records=int(match.group(1)), percent=100)
        return updated

    if updated.get("stage") == "discovering":
        target = max(1, updated.get("discovery_target", 1))
        updated["percent"] = min(25, round(updated.get("discovered", 0) / target * 25))
    elif updated.get("stage") in {"extracting", "enriching"}:
        total = max(1, updated.get("total", 1))
        current = updated.get("current", 0)
        updated["percent"] = min(99, 25 + round(current / total * 74))
    return updated


def read_results(path: Path, limit: int = 100, search: str = "") -> dict:
    limit = max(1, min(int(limit), 2000))
    columns: list[str] = []
    all_rows: list[dict] = []
    suffix = path.suffix.casefold()
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            headers = next(iterator, ())
            columns = [str(value or "") for value in headers]
            for values in iterator:
                all_rows.append({column: value for column, value in zip(columns, values)})
        finally:
            workbook.close()
    elif suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            columns = list(reader.fieldnames or [])
            all_rows = list(reader)
    elif suffix == ".json":
        raw = json.loads(path.read_text(encoding="utf-8"))
        all_rows = raw if isinstance(raw, list) else [raw]
        for row in all_rows:
            if isinstance(row, dict):
                for key in row:
                    if key not in columns:
                        columns.append(key)
        all_rows = [row for row in all_rows if isinstance(row, dict)]
    else:
        raise ValueError("Unsupported result format.")

    if search:
        needle = search.casefold()
        filtered = [row for row in all_rows if needle in " ".join(str(value or "") for value in row.values()).casefold()]
    else:
        filtered = all_rows
    return {"columns": columns, "rows": filtered[:limit], "total": len(all_rows), "filtered": len(filtered)}


class JobConflict(RuntimeError):
    pass


class JobManager:
    def __init__(self):
        self.lock = threading.RLock()
        self.jobs: dict[str, dict] = {}
        self.processes: dict[str, subprocess.Popen] = {}
        DATA_ROOT.mkdir(parents=True, exist_ok=True)
        OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
        self._load()

    def _load(self):
        if not HISTORY_FILE.exists():
            return
        try:
            raw = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
            for job in raw if isinstance(raw, list) else []:
                if job.get("status") in ACTIVE_STATUSES:
                    job["status"] = "interrupted"
                    job["finished_at"] = now_iso()
                    job.setdefault("logs", []).append("Dashboard restarted before this job finished.")
                if job.get("id"):
                    self.jobs[job["id"]] = job
        except (OSError, json.JSONDecodeError):
            self.jobs = {}

    def _persist(self):
        DATA_ROOT.mkdir(parents=True, exist_ok=True)
        serializable = []
        for job in sorted(self.jobs.values(), key=lambda item: item.get("created_at", "")):
            copy = deepcopy(job)
            copy["logs"] = copy.get("logs", [])[-MAX_LOG_LINES:]
            serializable.append(copy)
        temporary = HISTORY_FILE.with_suffix(".tmp")
        temporary.write_text(json.dumps(serializable, ensure_ascii=False, indent=2), encoding="utf-8")
        temporary.replace(HISTORY_FILE)

    def list(self) -> list[dict]:
        with self.lock:
            return [self._public(job, include_logs=False) for job in sorted(
                self.jobs.values(), key=lambda item: item.get("created_at", ""), reverse=True
            )]

    def get(self, job_id: str) -> dict | None:
        with self.lock:
            job = self.jobs.get(job_id)
            return self._public(job, include_logs=True) if job else None

    def _public(self, job: dict, include_logs: bool) -> dict:
        result = deepcopy(job)
        if not include_logs:
            result.pop("logs", None)
        result["download_ready"] = bool(result.get("output_file") and Path(result["output_file"]).is_file())
        return result

    def create(self, payload: dict) -> dict:
        config = validate_job_payload(payload)
        with self.lock:
            if any(job.get("status") in ACTIVE_STATUSES for job in self.jobs.values()):
                raise JobConflict("Another scrape is already running. Stop or finish it before starting a new job.")
            job_id = datetime.now().strftime("%Y%m%d-%H%M%S-") + uuid.uuid4().hex[:6]
            output_file = OUTPUT_ROOT / f"connected_businesses_{job_id}.{config['format']}"
            job = {
                "id": job_id,
                **config,
                "status": "queued",
                "created_at": now_iso(),
                "started_at": "",
                "finished_at": "",
                "output_file": str(output_file),
                "logs": ["Job queued by cloud dashboard."],
                "progress": {"stage": "queued", "percent": 0, "records": 0},
                "return_code": None,
                "stop_requested": False,
            }
            self.jobs[job_id] = job
            self._persist()
        threading.Thread(target=self._run, args=(job_id,), name=f"scrape-{job_id}", daemon=True).start()
        return self.get(job_id)

    def _append_log(self, job: dict, line: str):
        cleaned = line.rstrip()
        if not cleaned:
            return
        job.setdefault("logs", []).append(f"{datetime.now().strftime('%H:%M:%S')}  {cleaned}")
        job["logs"] = job["logs"][-MAX_LOG_LINES:]
        job["progress"] = parse_progress_line(cleaned, job.get("progress", {}))

    def _run(self, job_id: str):
        with self.lock:
            job = self.jobs[job_id]
            job["status"] = "running"
            job["started_at"] = now_iso()
            job["progress"] = {"stage": "starting", "percent": 1, "records": 0}
            self._persist()
            command = [
                sys.executable, "-u", str(PROJECT_ROOT / "connected_scraper.py"),
                "--query", job["query"], "--max-results", str(job["max_results"]),
                "--max-website-pages", str(job["max_website_pages"]),
                "--delay", str(job["delay"]), "--enrichment", job["enrichment"],
                "--format", job["format"], "--output", job["output_file"],
            ]
        process = None
        try:
            popen_options = {
                "cwd": str(PROJECT_ROOT), "stdout": subprocess.PIPE, "stderr": subprocess.STDOUT,
                "text": True, "encoding": "utf-8", "errors": "replace", "bufsize": 1,
            }
            if os.name == "nt":
                popen_options["creationflags"] = subprocess.CREATE_NEW_PROCESS_GROUP
            else:
                popen_options["start_new_session"] = True
            process = subprocess.Popen(command, **popen_options)
            with self.lock:
                self.processes[job_id] = process
                self._append_log(self.jobs[job_id], "Scraper process started.")
            assert process.stdout is not None
            for line in iter(process.stdout.readline, ""):
                with self.lock:
                    self._append_log(self.jobs[job_id], line)
                    self._persist()
            return_code = process.wait()
            with self.lock:
                job = self.jobs[job_id]
                job["return_code"] = return_code
                if job.get("stop_requested"):
                    job["status"] = "stopped"
                    job["progress"]["stage"] = "stopped"
                elif return_code == 0 and Path(job["output_file"]).is_file():
                    job["status"] = "completed"
                    job["progress"].update(stage="completed", percent=100)
                else:
                    job["status"] = "failed"
                    job["progress"]["stage"] = "failed"
                    self._append_log(job, f"Scraper exited with code {return_code}.")
                job["finished_at"] = now_iso()
                self._persist()
        except Exception as exc:
            with self.lock:
                job = self.jobs[job_id]
                job["status"] = "failed"
                job["finished_at"] = now_iso()
                job["progress"]["stage"] = "failed"
                self._append_log(job, f"Dashboard runner error: {exc}")
                self._persist()
        finally:
            with self.lock:
                self.processes.pop(job_id, None)

    def stop(self, job_id: str) -> dict | None:
        with self.lock:
            job = self.jobs.get(job_id)
            if not job:
                return None
            if job.get("status") not in ACTIVE_STATUSES:
                return self._public(job, include_logs=True)
            job["stop_requested"] = True
            job["status"] = "stopping"
            job["progress"]["stage"] = "stopping"
            self._append_log(job, "Stop requested. Saving collected checkpoint data...")
            process = self.processes.get(job_id)
            self._persist()
        if process and process.poll() is None:
            try:
                if os.name == "nt":
                    subprocess.run(
                        ["taskkill", "/PID", str(process.pid), "/T", "/F"],
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=10, check=False,
                    )
                else:
                    os.killpg(os.getpgid(process.pid), signal.SIGTERM)
            except Exception:
                process.terminate()
        return self.get(job_id)


MANAGER = JobManager()
AUTH = AuthManager()


class DashboardHandler(BaseHTTPRequestHandler):
    server_version = "ITCYBERScraperAPI/4.0"

    def log_message(self, format: str, *args):
        if getattr(self.server, "verbose", False):
            super().log_message(format, *args)

    def _cors_origin(self) -> str | None:
        origin = self.headers.get("Origin", "").strip().rstrip("/")
        return origin if origin and origin_is_allowed(origin) else None

    def _send_cors_headers(self):
        origin = self._cors_origin()
        if not origin:
            return
        # Echo the validated origin instead of sending "*" so browser
        # credentials/cookies can be used safely across Vercel -> Railway.
        self.send_header("Access-Control-Allow-Origin", origin)
        self.send_header("Access-Control-Allow-Credentials", "true")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Authorization, Content-Type")
        self.send_header("Access-Control-Expose-Headers", "Content-Disposition")
        self.send_header("Access-Control-Max-Age", "600")
        self.send_header("Vary", "Origin")

    def _reject_disallowed_origin(self) -> bool:
        origin = self.headers.get("Origin", "").strip()
        if origin and not origin_is_allowed(origin):
            self._json({"error": "This frontend origin is not allowed by the backend."}, HTTPStatus.FORBIDDEN)
            return True
        return False

    def _json(self, payload, status=HTTPStatus.OK):
        body = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "no-referrer")
        self._send_cors_headers()
        self.end_headers()
        self.wfile.write(body)

    def _session_token(self) -> str | None:
        authorization = self.headers.get("Authorization", "").strip()
        if authorization.lower().startswith("bearer "):
            token = authorization[7:].strip()
            return token if 20 <= len(token) <= 256 else None
        cookie = SimpleCookie()
        try:
            cookie.load(self.headers.get("Cookie", ""))
        except Exception:
            return None
        item = cookie.get(SESSION_COOKIE)
        return item.value if item else None

    def _request_uses_https_frontend(self) -> bool:
        origin = self.headers.get("Origin", "").strip().casefold()
        return origin.startswith("https://")

    def _set_session(self, token: str):
        cookie = SimpleCookie()
        cookie[SESSION_COOKIE] = token
        cookie[SESSION_COOKIE]["path"] = "/"
        cookie[SESSION_COOKIE]["httponly"] = True

        # Vercel frontend and Railway backend are different sites.
        # Cross-site cookies require SameSite=None + Secure.
        if self._request_uses_https_frontend():
            cookie[SESSION_COOKIE]["samesite"] = "None"
            cookie[SESSION_COOKIE]["secure"] = True
        else:
            # Keep localhost development working over plain HTTP.
            cookie[SESSION_COOKIE]["samesite"] = "Lax"

        cookie[SESSION_COOKIE]["max-age"] = str(SESSION_SECONDS)
        self.send_header("Set-Cookie", cookie.output(header="").strip())

    def _clear_session(self):
        cookie = SimpleCookie()
        cookie[SESSION_COOKIE] = ""
        cookie[SESSION_COOKIE]["path"] = "/"
        cookie[SESSION_COOKIE]["httponly"] = True

        if self._request_uses_https_frontend():
            cookie[SESSION_COOKIE]["samesite"] = "None"
            cookie[SESSION_COOKIE]["secure"] = True
        else:
            cookie[SESSION_COOKIE]["samesite"] = "Lax"

        cookie[SESSION_COOKIE]["max-age"] = "0"
        self.send_header("Set-Cookie", cookie.output(header="").strip())

    def _auth_user(self) -> dict | None:
        if AUTH_DISABLED:
            return dict(OPEN_DASHBOARD_USER)
        return AUTH.session_user(self._session_token())

    def _require_auth(self) -> dict | None:
        if AUTH_DISABLED:
            return dict(OPEN_DASHBOARD_USER)
        user = self._auth_user()
        if not user:
            self._json({"error": "Admin login required."}, HTTPStatus.UNAUTHORIZED)
        return user

    def _auth_response(self, token: str, user: dict, status=HTTPStatus.OK):
        body = json.dumps({
            "authenticated": True,
            "token": token,
            "expires_in": SESSION_SECONDS,
            "user": user,
        }, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "no-referrer")
        self._send_cors_headers()
        self._set_session(token)
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        if self._reject_disallowed_origin():
            return
        self.send_response(HTTPStatus.NO_CONTENT)
        self.send_header("Content-Length", "0")
        self.send_header("Cache-Control", "no-store")
        self._send_cors_headers()
        self.end_headers()

    def _read_json(self) -> dict:
        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError:
            raise ValueError("Invalid request length.")
        if length > 1_000_000:
            raise ValueError("Request is too large.")
        try:
            payload = json.loads(self.rfile.read(length) or b"{}")
        except json.JSONDecodeError as exc:
            raise ValueError("Invalid JSON request.") from exc
        if not isinstance(payload, dict):
            raise ValueError("Request body must be an object.")
        return payload

    def do_HEAD(self):
        path = urlparse(self.path).path.rstrip("/") or "/"
        if path in {"/", "/health", "/ready", "/api/health"}:
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Length", "0")
            self.send_header("Cache-Control", "no-store")
            self._send_cors_headers()
            self.end_headers()
            return
        self.send_response(HTTPStatus.NOT_FOUND)
        self.send_header("Content-Length", "0")
        self.end_headers()

    def do_GET(self):
        if self._reject_disallowed_origin():
            return
        parsed = urlparse(self.path)
        path = parsed.path.rstrip("/") or "/"
        if path == "/":
            return self._json({
                "service": "ITCYBER Connected Business Scraper API",
                "version": APP_VERSION,
                "status": "online",
                "dashboard": "Deploy the frontend folder to Vercel.",
            })
        if path in {"/health", "/ready"}:
            # Minimal unauthenticated endpoint for Railway health checks.
            # It intentionally avoids touching scraper/browser state.
            return self._json({"ok": True, "status": "healthy"})
        if path == "/api/health":
            return self._json({
                "ok": True,
                "version": APP_VERSION,
                "cloud_ready": True,
                "auth_required": False,
                "setup_required": False,
                "python": sys.version.split()[0],
                "scraper_ready": (PROJECT_ROOT / "connected_scraper.py").is_file(),
                "allowed_origins": list(configured_origins()),
            })
        if path == "/api/auth/status":
            user = self._auth_user()
            return self._json({
                "authenticated": True if AUTH_DISABLED else bool(user),
                "setup_required": False if AUTH_DISABLED else AUTH.setup_required,
                "auth_disabled": AUTH_DISABLED,
                "user": user,
            })
        if path.startswith("/api/") and not self._require_auth():
            return
        if path == "/api/jobs":
            return self._json({"jobs": MANAGER.list()})
        match = re.fullmatch(r"/api/jobs/([^/]+)", path)
        if match:
            job = MANAGER.get(match.group(1))
            return self._json(job if job else {"error": "Job not found."}, HTTPStatus.OK if job else HTTPStatus.NOT_FOUND)
        match = re.fullmatch(r"/api/jobs/([^/]+)/results", path)
        if match:
            job = MANAGER.get(match.group(1))
            if not job or not job.get("output_file") or not Path(job["output_file"]).is_file():
                return self._json({"error": "Result file is not ready."}, HTTPStatus.NOT_FOUND)
            params = parse_qs(parsed.query)
            try:
                result = read_results(
                    Path(job["output_file"]), int(params.get("limit", [100])[0]), params.get("q", [""])[0]
                )
                return self._json(result)
            except Exception as exc:
                return self._json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        match = re.fullmatch(r"/api/jobs/([^/]+)/download", path)
        if match:
            job = MANAGER.get(match.group(1))
            if not job or not job.get("output_file"):
                return self._json({"error": "Result file is not ready."}, HTTPStatus.NOT_FOUND)
            candidate = Path(job["output_file"]).resolve()
            try:
                candidate.relative_to(OUTPUT_ROOT.resolve())
            except ValueError:
                return self._json({"error": "Invalid output path."}, HTTPStatus.FORBIDDEN)
            if not candidate.is_file():
                return self._json({"error": "Result file is not ready."}, HTTPStatus.NOT_FOUND)
            body = candidate.read_bytes()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", mimetypes.guess_type(candidate.name)[0] or "application/octet-stream")
            self.send_header("Content-Disposition", f'attachment; filename="{candidate.name}"')
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.send_header("X-Content-Type-Options", "nosniff")
            self._send_cors_headers()
            self.end_headers()
            return self.wfile.write(body)
        return self._json({"error": "Endpoint not found."}, HTTPStatus.NOT_FOUND)

    def do_POST(self):
        if self._reject_disallowed_origin():
            return
        path = urlparse(self.path).path.rstrip("/")
        if AUTH_DISABLED and path.startswith("/api/auth/"):
            return self._json({
                "authenticated": True,
                "auth_disabled": True,
                "user": OPEN_DASHBOARD_USER,
                "message": "Authentication is disabled. Dashboard access is direct.",
            })
        if path == "/api/auth/setup":
            try:
                payload = self._read_json()
                if payload.get("password") != payload.get("confirm_password"):
                    raise ValueError("Password confirmation does not match.")
                token, user = AUTH.setup(payload.get("name", ""), payload.get("email", ""), payload.get("password", ""))
                return self._auth_response(token, user, HTTPStatus.CREATED)
            except RuntimeError as exc:
                return self._json({"error": str(exc)}, HTTPStatus.CONFLICT)
            except ValueError as exc:
                return self._json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        if path == "/api/auth/login":
            try:
                payload = self._read_json()
                token, user = AUTH.login(
                    payload.get("email", ""), payload.get("password", ""),
                    self.client_address[0] if self.client_address else "local",
                )
                return self._auth_response(token, user)
            except PermissionError as exc:
                return self._json({"error": str(exc)}, HTTPStatus.UNAUTHORIZED)
            except ValueError as exc:
                return self._json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        if path == "/api/auth/logout":
            AUTH.logout(self._session_token())
            body = json.dumps({"authenticated": False}).encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self._send_cors_headers()
            self._clear_session()
            self.end_headers()
            return self.wfile.write(body)
        if path == "/api/auth/change-password":
            user = self._require_auth()
            if not user:
                return
            try:
                payload = self._read_json()
                if payload.get("new_password") != payload.get("confirm_password"):
                    raise ValueError("New password confirmation does not match.")
                AUTH.change_password(
                    self._session_token() or "", payload.get("current_password", ""), payload.get("new_password", "")
                )
                return self._json({"ok": True, "message": "Password changed successfully."})
            except PermissionError as exc:
                return self._json({"error": str(exc)}, HTTPStatus.UNAUTHORIZED)
            except ValueError as exc:
                return self._json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        if not self._require_auth():
            return
        if path == "/api/jobs":
            try:
                job = MANAGER.create(self._read_json())
                return self._json(job, HTTPStatus.CREATED)
            except JobConflict as exc:
                return self._json({"error": str(exc)}, HTTPStatus.CONFLICT)
            except ValueError as exc:
                return self._json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        match = re.fullmatch(r"/api/jobs/([^/]+)/stop", path)
        if match:
            job = MANAGER.stop(match.group(1))
            return self._json(job if job else {"error": "Job not found."}, HTTPStatus.OK if job else HTTPStatus.NOT_FOUND)
        return self._json({"error": "Endpoint not found."}, HTTPStatus.NOT_FOUND)

def main() -> int:
    parser = argparse.ArgumentParser(description="Run the cloud-ready ITCYBER scraper API.")
    parser.add_argument("--host", default=os.environ.get("HOST", "0.0.0.0"))
    parser.add_argument("--port", type=int, default=int(os.environ.get("PORT", "8766")))
    parser.add_argument("--no-open", action="store_true", help="Do not open the browser automatically.")
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args()
    if not 1 <= args.port <= 65535:
        parser.error("--port must be between 1 and 65535")
    server = ThreadingHTTPServer((args.host, args.port), DashboardHandler)
    server.daemon_threads = True
    server.verbose = args.verbose
    url = f"http://{args.host}:{args.port}/"
    print(f"ITCYBER scraper API running at {url}", flush=True)
    print(f"Allowed frontend origins: {', '.join(configured_origins()) or '(none)'}", flush=True)
    # Do not try to open a browser on Railway/CI servers.
    running_in_cloud = bool(
        os.environ.get("RAILWAY_ENVIRONMENT")
        or os.environ.get("RAILWAY_PROJECT_ID")
        or os.environ.get("CI")
    )
    if not args.no_open and not running_in_cloud:
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nDashboard stopped.")
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
